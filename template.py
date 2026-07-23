"""
Exact Excel template builder for the OzMoEg Trip Planner.
Reproduces the structure, colours, dropdowns and formatting of
Desktop/Japan Trip.xlsx.
"""
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Side, Font, PatternFill, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# Colour palette (RGB hex as used by Excel)
HEADER_FILL = "002060"       # dark blue section headers
SUBHEADER_FILL = "B4C6E7"    # light blue title/weather/day rows
ITEM_FILL = "D8D8D8"         # light grey item rows column A
WHITE_FONT_THEME = 0
BLACK_FONT_THEME = 1
DEFAULT_FILL = "00000000"    # no fill (ARGB black/transparent)
GREEN_FILL = "92D050"        # bright green for selected schedule cells
HOTELS_HEADER_FILL = "0070C0"

THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
ALL_BORDERS = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


SCHEDULED_VALUES = {"Check in", "Check out", "Morning", "Afternoon", "Evening", "X", "Train", "Flight"}


def _fill(color: str, pattern: str = "solid") -> PatternFill:
    return PatternFill(start_color=color, end_color=color, patternType=pattern)


def _theme_color(theme: int):
    from openpyxl.styles.colors import Color
    return Color(theme=theme)


def _style_header_cell(cell, fill_color: str, font_theme: int = WHITE_FONT_THEME, bold: bool = False):
    cell.fill = _fill(fill_color)
    cell.font = Font(color=_theme_color(font_theme), bold=bold, name="Calibri", size=11)
    cell.alignment = CENTER_ALIGN
    cell.border = ALL_BORDERS


def _style_value_cell(cell, value=None, dark_blue_backup: bool = False, green: bool = False):
    cell.value = value
    cell.alignment = CENTER_ALIGN
    cell.border = ALL_BORDERS
    if dark_blue_backup:
        cell.fill = _fill(HEADER_FILL)
        cell.font = Font(name="Calibri", size=11, color=_theme_color(WHITE_FONT_THEME))
    elif green:
        cell.fill = _fill("92D050")
        cell.font = Font(name="Calibri", size=11, color=_theme_color(BLACK_FONT_THEME))
    else:
        cell.fill = PatternFill(fill_type=None)
        cell.font = Font(name="Calibri", size=11, color=_theme_color(BLACK_FONT_THEME))


SCHEDULED_VALUES_LOWER = {v.lower() for v in SCHEDULED_VALUES}


def build_template_workbook(trip_name: str, dates: List[datetime]) -> Workbook:
    """Create the skeleton workbook with headers, formatting and dropdowns."""
    wb = Workbook()
    ws = wb.active
    # Use the trip name as the sheet title; fall back to "Trip Plan" if empty.
    safe_title = trip_name[:31] if trip_name else "Trip Plan"
    ws.title = safe_title

    # Column widths
    ws.column_dimensions["A"].width = 62.43
    ws.column_dimensions["B"].width = 9.43
    ws.column_dimensions["C"].width = 14.0
    for col_idx in range(4, 4 + len(dates)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 17.0

    # Merge A1:A2
    ws.merge_cells("A1:A2")
    _style_header_cell(ws["A1"], SUBHEADER_FILL, WHITE_FONT_THEME)
    ws["A1"].value = trip_name

    # Row 1: dates in D onwards
    for i, d in enumerate(dates, start=4):
        cell = ws.cell(row=1, column=i)
        cell.value = d
        cell.number_format = "YYYY-MM-DD"
        _style_header_cell(cell, SUBHEADER_FILL, BLACK_FONT_THEME, bold=True)

    # Row 2: weekday names (literal values to avoid recalculation / zero bugs)
    for i, d in enumerate(dates, start=4):
        cell = ws.cell(row=2, column=i)
        cell.value = d.strftime("%A")
        _style_header_cell(cell, SUBHEADER_FILL, BLACK_FONT_THEME, bold=True)

    # Row 3: Weather
    _style_header_cell(ws["A3"], SUBHEADER_FILL, BLACK_FONT_THEME, bold=True)
    ws["A3"].value = "Weather"
    for i in range(4, 4 + len(dates)):
        cell = ws.cell(row=3, column=i)
        _style_header_cell(cell, SUBHEADER_FILL, BLACK_FONT_THEME, bold=True)

    # Row 4: day summary
    _style_header_cell(ws["A4"], SUBHEADER_FILL, BLACK_FONT_THEME, bold=True)
    ws["B4"].value = "Major city"
    _style_header_cell(ws["B4"], SUBHEADER_FILL, BLACK_FONT_THEME, bold=True)
    ws["C4"].value = "Distance from accommodation"
    _style_header_cell(ws["C4"], SUBHEADER_FILL, BLACK_FONT_THEME, bold=True)
    for i in range(4, 4 + len(dates)):
        cell = ws.cell(row=4, column=i)
        _style_header_cell(cell, SUBHEADER_FILL, BLACK_FONT_THEME, bold=True)

    # Freeze panes at D5
    ws.freeze_panes = "D5"

    return wb, ws


def add_section_header(ws, row: int, title: str, date_cols: int):
    cell = ws.cell(row=row, column=1)
    _style_header_cell(cell, HEADER_FILL, WHITE_FONT_THEME)
    cell.value = title
    # clear B/C to be safe and style dark-blue background across first 3 cols
    for c in range(2, 4):
        ws.cell(row=row, column=c).fill = _fill(HEADER_FILL)
        ws.cell(row=row, column=c).font = Font(color=_theme_color(WHITE_FONT_THEME))
        ws.cell(row=row, column=c).border = ALL_BORDERS
        ws.cell(row=row, column=c).alignment = CENTER_ALIGN
    for c in range(4, 4 + date_cols):
        _style_value_cell(ws.cell(row=row, column=c), dark_blue_backup=True)


def add_item_row(ws, row: int, name: str, major_city: str, distance: str,
                 date_cols: int):
    ws.cell(row=row, column=1).value = name
    ws.cell(row=row, column=1).fill = _fill(ITEM_FILL)
    ws.cell(row=row, column=1).font = Font(color=_theme_color(BLACK_FONT_THEME), name="Calibri", size=11)
    ws.cell(row=row, column=1).alignment = LEFT_ALIGN
    ws.cell(row=row, column=1).border = ALL_BORDERS

    ws.cell(row=row, column=2).value = major_city
    ws.cell(row=row, column=2).fill = _fill(ITEM_FILL)
    ws.cell(row=row, column=2).font = Font(color=_theme_color(BLACK_FONT_THEME), name="Calibri", size=11)
    ws.cell(row=row, column=2).alignment = CENTER_ALIGN
    ws.cell(row=row, column=2).border = ALL_BORDERS

    ws.cell(row=row, column=3).value = distance
    ws.cell(row=row, column=3).fill = _fill(ITEM_FILL)
    ws.cell(row=row, column=3).font = Font(color=_theme_color(BLACK_FONT_THEME), name="Calibri", size=11)
    ws.cell(row=row, column=3).alignment = CENTER_ALIGN
    ws.cell(row=row, column=3).border = ALL_BORDERS

    for c in range(4, 4 + date_cols):
        _style_value_cell(ws.cell(row=row, column=c), value="Backup")


def add_backup_rows(ws, start_row: int, date_cols: int, count: int = 3) -> int:
    """Add dark-blue empty placeholder rows at the bottom of the sheet."""
    row = start_row
    for _ in range(count):
        add_section_header(ws, row, "", date_cols)
        for c in range(4, 4 + date_cols):
            _style_value_cell(ws.cell(row=row, column=c), dark_blue_backup=True)
        row += 1
    return row


def add_hotel_transport_section(ws, start_row: int, cities: List[str],
                                transport_options: List[str],
                                date_cols: int,
                                transport_label: str = "Intercity Transfer") -> int:
    """Add Hotel/Transport header + hotel rows + transport row. Returns next row."""
    row = start_row
    add_section_header(ws, row, "Hotel/Transport", date_cols)
    row += 1
    hotel_start_row = row
    for city in cities:
        add_item_row(ws, row, city, "", "", date_cols)
        row += 1
    hotel_end_row = row - 1

    # Transport row(s)
    transport_row = row
    add_item_row(ws, row, transport_label, "", "", date_cols)
    row += 1

    # Dropdowns
    dv_time = DataValidation(type="list", formula1="\"Morning,Afternoon,Evening,X,Backup\"", allow_blank=True)
    ws.add_data_validation(dv_time)
    dv_time.add(f"D{start_row}:U{start_row}")

    dv_hotel = DataValidation(type="list", formula1="\"Check in, Check out,Backup\"", allow_blank=True)
    ws.add_data_validation(dv_hotel)
    dv_hotel.add(f"D{hotel_start_row}:U{hotel_end_row}")

    transport_choices = (transport_options or ["Train", "Flight", "Car", "Bus"]) + ["Backup"]
    dv_transport = DataValidation(type="list", formula1=f"\"{','.join(transport_choices)}\"", allow_blank=True)
    ws.add_data_validation(dv_transport)
    dv_transport.add(f"D{transport_row}:U{transport_row}")

    return row


def add_attraction_section(ws, start_row: int, title: str,
                           items: List[Dict], date_cols: int,
                           use_x_backup_only: bool = False) -> int:
    """Add a generic section header and item rows. Returns next row."""
    row = start_row
    add_section_header(ws, row, title, date_cols)
    header_row = row
    row += 1
    for item in items:
        add_item_row(ws, row, item["name"], item.get("city", ""),
                     item.get("distance", ""), date_cols)
        row += 1

    if use_x_backup_only:
        dv = DataValidation(type="list", formula1="\"X,Backup\"", allow_blank=True)
    else:
        dv = DataValidation(type="list", formula1="\"Morning,Afternoon,Evening,X,Backup\"", allow_blank=True)
    ws.add_data_validation(dv)
    last_item_row = row - 1
    dv.add(f"D{header_row}:U{last_item_row}")
    return row


def add_hotels_sheet(wb, hotels: List[Dict]):
    """Add Hotels tab matching the original format."""
    from openpyxl.styles.colors import Color
    ws = wb.create_sheet("Hotels")
    headers = ["City", "Hotel Name", "Dates", "Price (Approx.)", "Location", "Highlights", ""]
    ws.column_dimensions["A"].width = 7.71
    ws.column_dimensions["B"].width = 31.86
    ws.column_dimensions["C"].width = 10.29
    ws.column_dimensions["D"].width = 14.0
    ws.column_dimensions["E"].width = 12.71
    ws.column_dimensions["F"].width = 44.43
    ws.column_dimensions["G"].width = 13.0

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.value = h
        _style_header_cell(cell, HOTELS_HEADER_FILL, WHITE_FONT_THEME)

    row = 2
    current_city = None
    top_picked = False
    for idx, h in enumerate(hotels):
        is_city_header = h.get("city") and h["city"] != current_city
        if is_city_header:
            current_city = h["city"]
            cell = ws.cell(row=row, column=1)
            cell.value = current_city
            _style_header_cell(cell, HOTELS_HEADER_FILL, WHITE_FONT_THEME)
            top_picked = False
            for c in range(2, 8):
                _style_header_cell(ws.cell(row=row, column=c), HOTELS_HEADER_FILL, WHITE_FONT_THEME)
            row += 1

        # Mark first option per city as selected
        selected = not top_picked
        top_picked = True

        # If live data was missing, label it explicitly.
        price = h.get("price", "") if h.get("price") else "Failed to fetch"
        location = h.get("location", "") if h.get("location") else "Failed to fetch"
        highlights = h.get("highlights", "") if h.get("highlights") else "Failed to fetch"
        if isinstance(highlights, list):
            highlights = ", ".join(str(x) for x in highlights)
        elif not isinstance(highlights, str):
            highlights = str(highlights)

        values = [
            "",
            h["name"],
            h["dates"],
            price,
            location,
            highlights,
            "X" if selected else ""
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c)
            cell.value = v
            cell.border = ALL_BORDERS
            cell.alignment = LEFT_ALIGN if c in (2, 6) else CENTER_ALIGN
            cell.font = Font(name="Calibri", size=11, color=Color(theme=BLACK_FONT_THEME))
        row += 1


def build_skeleton(trip_name: str, dates: List[datetime],
                   cities: List[str], section_rows: Dict[str, List[Dict]],
                   transport_options: List[str],
                   hotels: List[Dict],
                   transport_label: str = "Intercity Transfer") -> Workbook:
    """High-level builder that creates the full formatted workbook skeleton."""
    wb, ws = build_template_workbook(trip_name, dates)
    date_cols = len(dates)

    # Hotel / Transport
    row = 5
    row = add_hotel_transport_section(ws, row, cities, transport_options, date_cols, transport_label)

    # Major Attractions
    row = add_attraction_section(ws, row, "Major Attractions & Landmarks",
                                 section_rows.get("attractions", []), date_cols)

    # Food & Dining
    row = add_attraction_section(ws, row, "Food & Dining Spots (Suggestions)",
                                 section_rows.get("food", []), date_cols)

    # Neighborhoods
    row = add_attraction_section(ws, row, "Neighborhoods & Areas",
                                 section_rows.get("neighborhoods", []), date_cols,
                                 use_x_backup_only=False)

    # Museums
    row = add_attraction_section(ws, row, "Museums & Cultural Stops",
                                 section_rows.get("museums", []), date_cols)

    # Markets
    row = add_attraction_section(ws, row, "Markets & Shopping",
                                 section_rows.get("markets", []), date_cols)

    # Day Trips
    row = add_attraction_section(ws, row, "Day Trip Destinations",
                                 section_rows.get("day_trips", []), date_cols)

    # Backup rows
    row = add_backup_rows(ws, row, date_cols)

    # Weather row fill
    for c in range(4, 4 + date_cols):
        ws.cell(row=3, column=c).fill = _fill(SUBHEADER_FILL)

    # AutoFilter on row 4 so users can filter each day's column (Backup vs Morning/Afternoon/Evening/X)
    last_col = get_column_letter(3 + date_cols)
    ws.auto_filter.ref = f"A4:{last_col}{row - 1}"

    add_hotels_sheet(wb, hotels)
    return wb


if __name__ == "__main__":
    # quick visual sanity test
    dates = [datetime(2025, 6, 12) + timedelta(days=i) for i in range(10)]
    cities = ["Tokyo", "Kyoto", "Osaka"]
    sections = {
        "attractions": [
            {"name": "Senso-ji Temple (Asakusa)", "city": "Tokyo"},
            {"name": "Ueno Park", "city": "Tokyo"},
        ],
        "food": [
            {"name": "Asakusa Imahan", "city": "Tokyo"},
        ],
        "neighborhoods": [
            {"name": "Shibuya", "city": "Tokyo"},
        ],
        "museums": [{"name": "Tokyo National Museum", "city": "Tokyo"}],
        "markets": [{"name": "Tsukiji Outer Market", "city": "Tokyo"}],
        "day_trips": [{"name": "Hakone", "city": "Tokyo"}],
    }
    hotels = [
        {"city": "Tokyo", "name": "Tokyu Stay Shinjuku", "dates": "13–17 June",
         "price": "AUD $220–240", "location": "Shinjuku",
         "highlights": "Kitchenette, washer/dryer"},
        {"city": "Tokyo", "name": "APA Hotel Shibuya", "dates": "13–17 June",
         "price": "AUD $180–200", "location": "Shibuya",
         "highlights": "Compact, clean, near nightlife"},
    ]
    wb = build_skeleton("Japan Trip", dates, cities, sections,
                        ["Train 2.5Hrs", "Train 30mins"], hotels)
    out = r"C:\Temp\trip_template_test.xlsx"
    wb.save(out)
    print("Saved skeleton to", out)
